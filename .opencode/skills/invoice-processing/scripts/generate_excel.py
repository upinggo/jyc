#!/usr/bin/env python3
"""Generate invoices.xlsx from SQLite database."""
import sys
import os
import shutil

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import load_workbook
from db import get_invoices_by_month

TEMPLATE_PATH = "template.xlsx"


def generate_excel(month: str, output_path: str = None) -> str:
    if not output_path:
        output_path = f"invoice_{month}/invoices.xlsx"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    shutil.copy(TEMPLATE_PATH, output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    invoices = get_invoices_by_month(month)

    for idx, inv in enumerate(invoices, start=2):
        ws.cell(row=idx, column=1, value=inv.get('seq') or idx - 1)
        ws.cell(row=idx, column=2, value=inv.get('invoice_no') or '')
        ws.cell(row=idx, column=3, value=inv.get('issue_date') or '')
        ws.cell(row=idx, column=4, value=inv.get('invoice_type') or '')
        ws.cell(row=idx, column=5, value=inv.get('buyer_name') or '')
        ws.cell(row=idx, column=6, value=inv.get('buyer_tax_id') or '')
        ws.cell(row=idx, column=7, value=inv.get('seller_name') or '')
        ws.cell(row=idx, column=8, value=inv.get('seller_tax_id') or '')
        ws.cell(row=idx, column=9, value=inv.get('service_name') or '')
        ws.cell(row=idx, column=10, value=inv.get('tax_rate') or '')
        ws.cell(row=idx, column=11, value=inv.get('amount'))
        ws.cell(row=idx, column=12, value=inv.get('tax'))
        ws.cell(row=idx, column=13, value=inv.get('total_amount'))
        ws.cell(row=idx, column=14, value=inv.get('remarks') or '')
        ws.cell(row=idx, column=15, value=inv.get('filename') or '')

    wb.save(output_path)
    wb.close()
    return output_path


if __name__ == "__main__":
    month = sys.argv[1] if len(sys.argv) > 1 else None
    if not month:
        print("Usage: python generate_excel.py <month>")
        sys.exit(1)

    output = generate_excel(month)
    print(f"Generated: {output}")