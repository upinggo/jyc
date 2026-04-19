#!/usr/bin/env python3
"""Generate invoice_summary_<month>.xlsx from SQLite database."""
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import load_workbook
from db import get_invoices_by_month

TEMPLATE_PATH = os.environ.get('INVOICE_SUMMARY_TEMPLATE_PATH', 'summary.xlsx')
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(SCRIPT_DIR)

MONTH_MAP = {
    '01': 'January', '02': 'February', '03': 'March', '04': 'April',
    '05': 'May', '06': 'June', '07': 'July', '08': 'August',
    '09': 'September', '10': 'October', '11': 'November', '12': 'December'
}

CATEGORY_KEYWORDS = {
    10: ['洗衣', '干洗', 'laundry'],
    12: ['餐饮', '餐费', '食品', '外卖', 'food', 'meal'],
    16: ['机票', '航空', 'airticket', 'flight'],
}


def month_to_str(month: str) -> str:
    """Convert YYYY-MM to 'Month of YYYY' format."""
    try:
        year, mon = month.split('-')
        return f"{MONTH_MAP.get(mon, mon)} of {year}"
    except:
        return month


def generate_summary_excel(month: str, output_path: str = None) -> str:
    template_path = TEMPLATE_PATH
    if not os.path.isabs(template_path):
        template_path = os.path.join(os.getcwd(), template_path)

    if not os.path.exists(template_path):
        fallback_path = os.path.join(SKILL_DIR, 'summary.xlsx')
        if os.path.exists(fallback_path):
            template_path = fallback_path
        else:
            raise FileNotFoundError(
                f"Summary template not found: {template_path}. "
                "Please ensure summary.xlsx exists in the working directory "
                "or set INVOICE_SUMMARY_TEMPLATE_PATH environment variable."
            )

    if not output_path:
        output_path = f"invoice_summary_{month}.xlsx"

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    shutil.copy(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    ws['B6'] = month_to_str(month)

    invoices = get_invoices_by_month(month)

    row_totals = {row: 0.0 for row in CATEGORY_KEYWORDS}

    for inv in invoices:
        service = str(inv.get('service_name') or '').lower()
        date = inv.get('issue_date')
        amount = inv.get('total_amount')

        if not amount:
            continue

        target_row = None
        for cat_row, keywords in CATEGORY_KEYWORDS.items():
            if any(kw.lower() in service for kw in keywords):
                target_row = cat_row
                break

        if target_row:
            row_totals[target_row] += float(amount)
            if not ws.cell(row=target_row, column=2).value:
                ws.cell(row=target_row, column=2, value=date)

    for row, total in row_totals.items():
        if total > 0:
            ws.cell(row=row, column=3, value=total)

    wb.save(output_path)
    wb.close()
    return output_path


if __name__ == "__main__":
    month = sys.argv[1] if len(sys.argv) > 1 else None
    if not month:
        print("Usage: python generate_summary_excel.py <month>")
        sys.exit(1)

    output = generate_summary_excel(month)
    print(f"Generated: {output} (monthly summary)")