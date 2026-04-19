#!/usr/bin/env python3
"""Migrate existing Excel invoices to SQLite database."""
import sys
import os
import glob
import re

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import load_workbook
from db import init_db, check_duplicate, insert_invoice


def migrate_excel_to_db(excel_path: str, month: str) -> dict:
    result = {'success': 0, 'skipped': 0, 'failed': 0, 'errors': []}
    if not os.path.exists(excel_path):
        result['errors'].append(f"File not found: {excel_path}")
        return result

    wb = load_workbook(excel_path)
    ws = wb.active

    for row_num in range(2, ws.max_row + 1):
        try:
            invoice_no = str(ws.cell(row=row_num, column=2).value or '').strip()
            seller_tax_id = str(ws.cell(row=row_num, column=8).value or '').strip()
            total_amount = ws.cell(row=row_num, column=13).value

            if not seller_tax_id and not total_amount:
                continue
            if not seller_tax_id or len(seller_tax_id) != 18:
                result['skipped'] += 1
                continue
            if not total_amount:
                result['skipped'] += 1
                continue

            total_amount = float(str(total_amount).replace(',', ''))

            existing = check_duplicate(invoice_no, seller_tax_id, total_amount)
            if existing:
                result['skipped'] += 1
                continue

            remarks = str(ws.cell(row=row_num, column=14).value or '')
            verify_code = ''
            if '校验码:' in remarks:
                match = re.search(r'校验码:\s*(\d{20})', remarks)
                if match:
                    verify_code = match.group(1)

            data = {
                'seq': ws.cell(row=row_num, column=1).value,
                'invoice_no': invoice_no,
                'issue_date': str(ws.cell(row=row_num, column=3).value or ''),
                'invoice_type': str(ws.cell(row=row_num, column=4).value or ''),
                'buyer_name': str(ws.cell(row=row_num, column=5).value or ''),
                'buyer_tax_id': str(ws.cell(row=row_num, column=6).value or ''),
                'seller_name': str(ws.cell(row=row_num, column=7).value or ''),
                'seller_tax_id': seller_tax_id,
                'service_name': str(ws.cell(row=row_num, column=9).value or ''),
                'tax_rate': str(ws.cell(row=row_num, column=10).value or ''),
                'amount': ws.cell(row=row_num, column=11).value,
                'tax': ws.cell(row=row_num, column=12).value,
                'total_amount': total_amount,
                'remarks': remarks,
                'filename': str(ws.cell(row=row_num, column=15).value or ''),
                'verify_code': verify_code,
                'source': 'migration',
                'month': month,
            }

            insert_invoice(data)
            result['success'] += 1

        except Exception as e:
            result['failed'] += 1
            result['errors'].append(f"Row {row_num}: {str(e)}")

    wb.close()
    return result


def main():
    init_db()
    print("=== Invoice Migration Tool ===\n")

    excel_files = glob.glob("invoice_list_*.xlsx")
    if not excel_files:
        print("No invoice_list_*.xlsx files found.")
        return

    total = {'success': 0, 'skipped': 0, 'failed': 0}

    for excel_path in sorted(excel_files):
        month = excel_path.replace('invoice_list_', '').replace('.xlsx', '')
        print(f"\nProcessing {excel_path} (month: {month})...")

        result = migrate_excel_to_db(excel_path, month)
        print(f"  Success: {result['success']}, Skipped: {result['skipped']}, Failed: {result['failed']}")

        total['success'] += result['success']
        total['skipped'] += result['skipped']
        total['failed'] += result['failed']

    print(f"\n=== Summary ===")
    print(f"Total: success={total['success']}, skipped={total['skipped']}, failed={total['failed']}")


if __name__ == "__main__":
    main()
